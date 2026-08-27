# -*- coding: utf-8 -*-
"""เขียนรายชื่อสัญญาที่ถึงเกณฑ์บอกเลิกสัญญาลงชีตใหม่ในไฟล์ Excel เดิม

- Sheet1 (ตัวอย่างของทีม) ไม่ถูกแตะ
- สร้างชีตใหม่ชื่อ 'OD6 <เดือน>' โครงคอลัมน์ A–AG เหมือน Sheet1 ทุกช่อง
  รวมสูตร =BAHTTEXT() 7 ช่อง  แล้วต่อฟิลด์เสริมที่คอลัมน์ AH เป็นต้นไป

รันด้วย:  $env:K2_USER="..."; $env:K2_PWD="..."; python k2_termination_to_excel.py
"""
import os
import shutil
import sys
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pyodbc

XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                    "ตัวอย่างข้อมูลที่ใช้ในหนังสือบอกเลิกสัญญา.xlsx")
SQL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "k2_termination_list.sql")
SHEET = "OD6 ส.ค.2569"

USERNAME = os.environ.get("K2_USER")
PASSWORD = os.environ.get("K2_PWD")
if not USERNAME or not PASSWORD:
    sys.exit('ต้องตั้ง credential ก่อน:  $env:K2_USER="..."; $env:K2_PWD="..."')

# คอลัมน์ A–AG ของ Sheet1 · None = ช่องสูตร BAHTTEXT
# (ตำแหน่ง Excel, ชื่อคอลัมน์ที่ SQL คืนมา)
LETTER_COLS = [
    (1,  "A_สัญญาเลขที่"),        (2,  "B_ชื่อผู้ทำสัญญา"),   (3,  "C_วันที่ทำสัญญา"),
    (4,  "D_ประเภทสัญญา"),        (5,  "E_ยี่ห้อ"),           (6,  "F_รุ่นแบบ"),
    (7,  "G_เครื่อง"),            (8,  "H_SerialNumber"),     (9,  "I_ค่าเช่าซื้อรวมVAT"),
    (10, None),                    (11, "K_งวดละ"),            (12, None),
    (13, "M_จำนวนงวด"),           (14, "N_วันเริ่มชำระงวดแรก"), (15, "O_วันที่ชำระงวดต่อไป"),
    (16, "P_ค้างชำระงวดแรก"),      (17, "Q_งวดที่ครบOD6"),      (18, "R_ค่าเช่าซื้อค้าง"),
    (19, None),                    (20, "T_ค่าเบี้ยปรับ"),      (21, None),
    (22, "V_ค่าติดตามทวงถาม"),     (23, None),                 (24, "X_ค่าใช้จ่ายอื่น"),
    (25, None),                    (26, "Z_รวมเป็นเงินทั้งสิ้น"), (27, None),
    (28, "AB_วันที่เป็นOD6"),      (29, "AC_ที่อยู่ตามทะเบียนบ้าน"), (30, "AD_เลขที่"),
    (31, "AE_ตำบลอำเภอ"),          (32, "AF_จังหวัด"),          (33, "AG_รหัสไปรษณีย์"),
]
BAHTTEXT_SRC = {10: "I", 12: "K", 19: "R", 21: "T", 23: "V", 25: "X", 27: "Z"}


def main():
    conn = pyodbc.connect(
        "DRIVER={ODBC Driver 18 for SQL Server};SERVER=43.254.133.123;DATABASE=HPCOM7;"
        f"UID={USERNAME};PWD={PASSWORD};TrustServerCertificate=yes;", timeout=900)
    cur = conn.cursor()
    print("กำลังดึงข้อมูล ...")
    cur.execute(open(SQL_FILE, encoding="utf-8").read())
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    conn.close()
    print(f"  ได้ {len(rows):,} สัญญา · {len(cols)} คอลัมน์")

    letter_names = {n for _, n in LETTER_COLS if n}
    extra = [c for c in cols if c not in letter_names]

    shutil.copy2(XLSX, XLSX.replace(".xlsx", " (backup ก่อนลงรายชื่อ).xlsx"))
    wb = openpyxl.load_workbook(XLSX)
    src = wb["Sheet1"]
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    extra_fill = PatternFill("solid", fgColor="FCE4D6")
    bold = Font(bold=True)

    # แถว 1 = ป้ายกลุ่ม · แถว 2 = ชื่อคอลัมน์ (ให้เหมือน Sheet1 ที่ header อยู่แถว 2)
    ws.cell(1, 1, "ช่องที่ใช้ในหนังสือ (A–AG) — เหมือน Sheet1").font = bold
    ws.cell(1, 34, "ฟิลด์เสริม — ไม่ได้อยู่ในหนังสือ").font = bold

    for pos, name in LETTER_COLS:
        h = src.cell(2, pos).value
        c = ws.cell(2, pos, h)
        c.font, c.fill = bold, hdr_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for j, name in enumerate(extra):
        c = ws.cell(2, 34 + j, name)
        c.font, c.fill = bold, extra_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")

    for i, r in enumerate(rows):
        er = 3 + i
        for pos, name in LETTER_COLS:
            cell = ws.cell(er, pos)
            if name is None:
                cell.value = f"=BAHTTEXT({BAHTTEXT_SRC[pos]}{er})"
            else:
                v = r[name]
                cell.value = datetime.combine(v, datetime.min.time()) if isinstance(v, date) and not isinstance(v, datetime) else v
            cell.number_format = src.cell(3, pos).number_format
        for j, name in enumerate(extra):
            v = r[name]
            ws.cell(er, 34 + j,
                    datetime.combine(v, datetime.min.time())
                    if isinstance(v, date) and not isinstance(v, datetime) else v)

    ws.freeze_panes = "B3"
    for pos in range(1, 34 + len(extra)):
        ws.column_dimensions[get_column_letter(pos)].width = 18
    ws.column_dimensions["B"].width = 26
    for L in ("AC", "AD"):
        ws.column_dimensions[L].width = 45

    wb.save(XLSX)

    from collections import Counter
    print(f"\nเขียนลงชีต '{SHEET}' แถว 3–{2+len(rows)}")
    print("แยกตามเกณฑ์:", dict(Counter(r["เกณฑ์"] for r in rows)))
    print(f"ยอดรวมจากการ์ด {sum(r['Z_รวมเป็นเงินทั้งสิ้น'] for r in rows):,.2f} บาท")
    print(f"ช่องในหนังสือ 33 (รวมสูตร BAHTTEXT 7 ช่อง) · ฟิลด์เสริม {len(extra)} ช่อง")


if __name__ == "__main__":
    main()
