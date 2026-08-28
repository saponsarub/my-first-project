# -*- coding: utf-8 -*-
"""รัน sql/k2_termination_list_v2.sql แล้วเขียนเป็น Excel ตาม Template จดหมายบอกเลิก

    $env:K2_USER="..."; $env:K2_PWD="..."
    python scripts/k2/k2_termination_v2_to_excel.py [outfile.xlsx]

ผลลัพธ์ 2 sheet
  ข้อมูล   — ตรงตาม template (มีคอลัมน์ตัวหนังสือเป็นสูตร =BAHTTEXT)
  ตรวจสอบ  — คอลัมน์ _* ไว้ตรวจก่อนออกหนังสือ

⚠️ ไฟล์ผลลัพธ์มี PII — ห้าม commit เข้า git (.gitignore กัน *.xlsx ไว้แล้ว)
"""
import os
import sys
import pathlib

import pyodbc
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SQL = pathlib.Path(__file__).resolve().parents[2] / "sql" / "k2_termination_list_v2.sql"
OUT = pathlib.Path(sys.argv[1] if len(sys.argv) > 1 else "k2_termination_v2.xlsx")

USER, PWD = os.environ.get("K2_USER"), os.environ.get("K2_PWD")
if not USER or not PWD:
    sys.exit('ต้องตั้ง credential ก่อน:  $env:K2_USER="..."; $env:K2_PWD="..."')

cn = pyodbc.connect(
    "DRIVER={ODBC Driver 18 for SQL Server};SERVER=43.254.133.123;DATABASE=HPCOM7;"
    f"UID={USER};PWD={PWD};TrustServerCertificate=yes;", timeout=600)
cur = cn.cursor()
cur.execute(SQL.read_text(encoding="utf-8"))
cols = [d[0] for d in cur.description]
rows = [list(r) for r in cur.fetchall()]
cn.close()
print(f"ดึงมา {len(rows)} แถว · {len(cols)} คอลัมน์")

# คอลัมน์เงินที่ต้องมีคู่ตัวหนังสือ (=BAHTTEXT)
MONEY = ["ค่าเช่าซื้อรวมภาษีมูลค่าเพิ่ม", "งวดละ (บาท)", "ค่าเช่าซื้อ 6 งวด (OD1-OD6)",
         "ค่าเบี้ยปรับชำระล่าช้า", "ค่าติดตามทวงถามหนี้",
         "ค่าใช้จ่ายอื่นที่บริษัทฯ มีสิทธิเรียกเก็บ", "รวมเป็นเงินทั้งสิ้น"]

letter_cols = [c for c in cols if not c.startswith("_")]
check_cols = ["สัญญาเลขที่"] + [c for c in cols if c.startswith("_")]

wb = openpyxl.Workbook()
HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="4472C4")
FILL2 = PatternFill("solid", fgColor="ED7D31")


def write_sheet(ws, headers, with_bahttext):
    out_hdr = []
    for h in headers:
        out_hdr.append(h)
        if with_bahttext and h in MONEY:
            out_hdr.append(h + " (ตัวหนังสือ)")
    ws.append(out_hdr)
    for c in range(1, len(out_hdr) + 1):
        cell = ws.cell(1, c)
        cell.font = HDR
        cell.fill = FILL if not with_bahttext else FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"

    for ri, r in enumerate(rows, start=2):
        ci = 1
        for h in headers:
            v = r[cols.index(h)]
            ws.cell(ri, ci, v)
            if with_bahttext and h in MONEY:
                num_letter = get_column_letter(ci)
                ws.cell(ri, ci + 1, f"=BAHTTEXT({num_letter}{ri})")
                ci += 1
            ci += 1
    for c in range(1, len(out_hdr) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22


ws1 = wb.active
ws1.title = "ข้อมูล"
write_sheet(ws1, letter_cols, True)

ws2 = wb.create_sheet("ตรวจสอบ")
write_sheet(ws2, check_cols, False)
for c in range(1, len(check_cols) + 1):
    ws2.cell(1, c).fill = FILL2

wb.save(OUT)
print(f"เขียน {OUT.resolve()}")
print("⚠️ ไฟล์นี้มี PII — ห้าม commit เข้า git")
print("⚠️ คอลัมน์ตัวหนังสือเป็นสูตร =BAHTTEXT ต้องเปิดด้วย Excel เท่านั้นถึงจะได้ค่า")
