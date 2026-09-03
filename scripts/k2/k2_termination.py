# -*- coding: utf-8 -*-
"""ทำรายชื่อ + Excel หนังสือบอกเลิกสัญญา — เรียกใช้ได้ทุกเดือน ทุกวัน

    $env:K2_USER="..."; $env:K2_PWD="..."

    python scripts/k2/k2_termination.py                # ถึงวันนี้ · ทุกรอบ
    python scripts/k2/k2_termination.py 2026-09        # เดือน ก.ย. 2026 ทั้ง 2 รอบ
    python scripts/k2/k2_termination.py 2026-09-01     # เฉพาะรอบวันที่ 1
    python scripts/k2/k2_termination.py 2026-09-16     # เฉพาะรอบวันที่ 16
    python scripts/k2/k2_termination.py 9/16           # ปีปัจจุบัน รอบวันที่ 16

ตัวเลือก
    -o FILE            ชื่อไฟล์ผลลัพธ์ (ปกติตั้งให้อัตโนมัติ)
    --asof DATE        นับงวดค้างถึงวันไหน (ปกติ = วันนี้ หรือสิ้นเดือนที่ระบุ)
    --min-od N         ต้องค้างอย่างน้อยกี่งวดจึงเข้ารายชื่อ (ปกติ 6)
    --all-od6          เอาทุกคนที่ค้างครบเกณฑ์ ไม่ว่าจะครบ OD6 เมื่อไหร่
    --template PATH    ไฟล์ template (ปกติอ่านจาก env TERMINATION_TEMPLATE)
    --csv              เขียน CSV เพิ่มอีก 1 ไฟล์
    --dry-run          แค่นับจำนวน ไม่เขียนไฟล์

วิธีคัด (v4 · 2026-08-31)
    คัดจาก CUSTOMER_CARD ล้วน — นับงวดที่ RECEIPT_NUMBER IS NULL และ DUEDATE <= asof
    เลิกใช้ COLLECTION_OD.CONTRACT_STATUS = 48 เพราะ snapshot อัปเดตช้าและป้ายเหนียว
    ทำให้รอบ 8/2026 มี 204 จาก 617 รายที่ค้างจริงไม่ถึง 6 งวด

    การ์ดอัปเดตทันทีที่รับชำระ รันวันไหนก็ได้ภาพของวันนั้นเสมอ
    ไม่ต้องรอวัน refresh ของ snapshot อีกต่อไป

    เอาเฉพาะคนที่ **ครบ OD6 ในรอบเดือนที่ระบุ** — คนที่ครบตั้งแต่เดือนก่อน
    ต้องอยู่ในรายชื่อของเดือนนั้นไปแล้ว ส่งซ้ำไม่ได้ (ปิดด้วย --all-od6)

    สัญญาที่จบไปแล้วนับงวดค้างเป็น "ในสัญญา + นอกสัญญา" ถึง asof
    เพราะการ์ดหยุดออกงวดตั้งแต่วันสิ้นสุดสัญญา ถ้านับแค่งวดว่างจะไม่มีวันถึง 6

⚠️ ไฟล์ผลลัพธ์มี PII — ห้าม commit เข้า git
"""
import argparse
import calendar
import datetime as dt
import os
import pathlib
import re
import sys
from copy import copy

import pyodbc
import openpyxl
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

ROOT = pathlib.Path(__file__).resolve().parents[2]
SQL_FILE = ROOT / "sql" / "k2_termination_list_v4.sql"
BE = 543                      # พ.ศ. − ค.ศ.
MIN_OD_DEFAULT = 6            # ค้างครบ 6 งวดจึงเกิดสิทธิ์บอกเลิก

# ---- คอลัมน์ใน template 'ข้อมูล' → คอลัมน์จากผลลัพธ์ SQL ------------------
MAP = [
    ("A",  "สัญญาเลขที่",                          "str"),
    ("B",  "ชื่อผู้ทำสัญญา",                        "str"),
    ("C",  "วันที่ทำสัญญา",                         "date"),
    ("D",  "ประเภทสัญญา",                          "str"),
    ("E",  "ยี่ห้อ",                                "str"),
    ("F",  "รุ่น/แบบ",                              "str"),
    ("G",  "เครื่อง",                               "int"),
    ("H",  "Serial Number",                        "str"),
    ("I",  "ค่าเช่าซื้อรวมภาษีมูลค่าเพิ่ม",          "float"),
    ("J",  "I",                                    "bahttext"),
    ("K",  "งวดละ (บาท)",                          "float"),
    ("L",  "K",                                    "bahttext"),
    ("M",  "จำนวนงวดเช่าซื้อ",                      "int"),
    ("N",  "วันเริ่มชำระงวดแรก",                    "date"),
    ("O",  "วันที่ชำระงวดต่อไป",                    "int"),
    ("P",  "ค้างชำระงวดแรก",                        "int"),
    ("Q",  "งวดที่ครบ OD 6",                        "int"),
    ("R",  "ค่าเช่าซื้อที่ค้างชำระ",                "float"),
    ("S",  "R",                                    "bahttext"),
    ("T",  "ค่าเบี้ยปรับชำระล่าช้า",                "float"),
    ("U",  "T",                                    "bahttext"),
    ("V",  "ค่าติดตามทวงถามหนี้",                   "float"),
    ("W",  "V",                                    "bahttext"),
    ("X",  "ค่าใช้จ่ายอื่นที่บริษัทฯ มีสิทธิเรียกเก็บ", "float"),
    ("Y",  "X",                                    "bahttext"),
    ("Z",  "รวมเป็นเงินทั้งสิ้น",                    "float"),
    ("AA", "Z",                                    "bahttext"),
    ("AB", "วันที่เป็น OD6",                        "dateBE"),
    ("AC", "รวมจำนวนงวดที่ค้าง",                    "int"),
    ("AD", "__ADDRESS_FULL__",                     "str"),
    ("AE", "เลขที่",                                "str"),
    ("AF", "แขวง/ตำบล/อำเภอ",                      "str"),
    ("AG", "จังหวัด",                               "str"),
    ("AH", "รหัส ปณ.",                              "int"),
]


# ---------------------------------------------------------------- args ----
def parse_period(text, today):
    """คืน (year, month, rounds) โดย rounds เป็น subset ของ {1, 16}"""
    if not text:
        # ไม่ระบุ → เดือนปัจจุบัน · เลือกรอบจากวันที่วันนี้
        rounds = {1} if today.day < 16 + 2 else {1, 16}
        return today.year, today.month, rounds

    t = text.strip().replace("/", "-")
    parts = [p for p in t.split("-") if p]
    nums = [int(p) for p in parts]

    if len(nums) == 1:                       # "9"
        return today.year, nums[0], {1, 16}
    if len(nums) == 2:
        a, b = nums
        if a > 31:                           # "2026-09"
            return a, b, {1, 16}
        if b in (1, 16):                     # "9-16"
            return today.year, a, {b}
        raise SystemExit(f"รอบต้องเป็นวันที่ 1 หรือ 16 เท่านั้น (ได้ {b})")
    if len(nums) == 3:                       # "2026-09-16"
        y, m, d = nums
        if d not in (1, 16):
            raise SystemExit(f"รอบต้องเป็นวันที่ 1 หรือ 16 เท่านั้น (ได้ {d})")
        return y, m, {d}
    raise SystemExit(f"อ่านรอบไม่ออก: {text!r}")


ap = argparse.ArgumentParser(add_help=True)
ap.add_argument("period", nargs="?", help="เช่น 2026-09 · 2026-09-16 · 9/16")
ap.add_argument("-o", "--out")
ap.add_argument("--asof", help="นับงวดค้างถึงวันไหน (ปกติ = วันนี้)")
ap.add_argument("--min-od", type=int, default=MIN_OD_DEFAULT,
                help=f"ต้องค้างอย่างน้อยกี่งวด (ปกติ {MIN_OD_DEFAULT})")
ap.add_argument("--all-od6", action="store_true",
                help="เอาทุกคนที่ค้างครบเกณฑ์ ไม่ว่าจะครบเมื่อไหร่ "
                     "(ปกติเอาเฉพาะคนที่เพิ่งครบในรอบเดือนนั้น)")
ap.add_argument("--template")
ap.add_argument("--csv", action="store_true")
ap.add_argument("--dry-run", action="store_true")
args = ap.parse_args()

TODAY = dt.date.today()
YEAR, MONTH, ROUNDS = parse_period(args.period, TODAY)
if not 1 <= MONTH <= 12:
    raise SystemExit(f"เดือนไม่ถูกต้อง: {MONTH}")

TEMPLATE = pathlib.Path(args.template or os.environ.get(
    "TERMINATION_TEMPLATE",
    r"C:\Users\Sapon.S\Downloads\Template จดหมายบอกเลิก.xlsx"))

USER, PWD = os.environ.get("K2_USER"), os.environ.get("K2_PWD")
if not USER or not PWD:
    raise SystemExit('ต้องตั้ง credential ก่อน:  $env:K2_USER="..."; $env:K2_PWD="..."')

if args.min_od < 1:
    raise SystemExit("--min-od ต้องเป็นจำนวนเต็มบวก")

cut = dt.date(YEAR, MONTH, 1)                      # เส้นแบ่งกรณี 1 / กรณี 2
eom = dt.date(YEAR, MONTH, calendar.monthrange(YEAR, MONTH)[1])
asof = dt.date.fromisoformat(args.asof) if args.asof else min(TODAY, eom)
if asof < cut:
    raise SystemExit(f"--asof ({asof}) ต้องไม่ก่อนต้นเดือนที่ระบุ ({cut})")
rnd_param = 0 if ROUNDS == {1, 16} else sorted(ROUNDS)[0]

label = f"{YEAR}-{MONTH:02d}"
print(f"รอบ {label} · {' + '.join(f'วันที่ {r}' for r in sorted(ROUNDS))}")
print(f"  นับงวดค้างจาก CUSTOMER_CARD ถึง {asof} · ต้องค้าง >= {args.min_od} งวด")
print("  " + ("เอาทุกคนที่ค้างครบเกณฑ์ ไม่ว่าจะครบเมื่อไหร่"
              if args.all_od6 else f"เอาเฉพาะคนที่ครบ OD6 ในเดือน {label}"))

# ------------------------------------------------------------ connect ----
cn = pyodbc.connect(
    "DRIVER={ODBC Driver 18 for SQL Server};SERVER=43.254.133.123;DATABASE=HPCOM7;"
    f"UID={USER};PWD={PWD};TrustServerCertificate=yes;", timeout=900)
cur = cn.cursor()


# --------------------------------------------------------------- SQL ----
sql = SQL_FILE.read_text(encoding="utf-8")

dates = {"@asof": asof, "@cut": cut}
ints = {"@minod": args.min_od, "@round": rnd_param,
        "@newonly": 0 if args.all_od6 else 1}

for var, val in dates.items():
    sql, n = re.subn(rf"(DECLARE\s+{re.escape(var)}\s+date\s*=\s*)'[^']*'",
                     lambda m: m.group(1) + f"'{val}'", sql)
    if n != 1:
        cn.close()
        raise SystemExit(f"แทนค่า {var} ใน SQL ไม่ได้ (เจอ {n} จุด)")

for var, val in ints.items():
    sql, n = re.subn(rf"(DECLARE\s+{re.escape(var)}\s+int\s*=\s*)-?\d+",
                     lambda m: m.group(1) + str(val), sql)
    if n != 1:
        cn.close()
        raise SystemExit(f"แทนค่า {var} ใน SQL ไม่ได้ (เจอ {n} จุด)")

cur.execute(sql)
cols = [d[0] for d in cur.description]
rows = [list(r) for r in cur.fetchall()]
cn.close()

ix = {c: i for i, c in enumerate(cols)}
n1 = sum(1 for r in rows if r[ix["กรณี"]] == 1)
n2 = sum(1 for r in rows if r[ix["กรณี"]] == 2)
print(f"\nได้ {len(rows)} สัญญา · กรณี 1 (OD6 ปกติ) = {n1} · กรณี 2 (สัญญาหมดอายุ) = {n2}")
if not rows:
    raise SystemExit("ไม่มีข้อมูล — ไม่เขียนไฟล์")

import collections
dist = collections.Counter(r[ix["รวมจำนวนงวดที่ค้าง"]] for r in rows)
print("  งวดที่ค้าง: " + " · ".join(f"{k} งวด = {dist[k]}" for k in sorted(dist)))
print(f"  ยอดเรียกเก็บรวม {sum(r[ix['รวมเป็นเงินทั้งสิ้น']] or 0 for r in rows):,.2f} บาท")

if args.dry_run:
    raise SystemExit(0)

# -------------------------------------------------------------- Excel ----
if not TEMPLATE.exists():
    raise SystemExit(f"ไม่พบ template: {TEMPLATE}")

rnd = "" if ROUNDS == {1, 16} else f" รอบ{sorted(ROUNDS)[0]}"
out = pathlib.Path(args.out or f"หนังสือบอกเลิกสัญญา {label}{rnd} ณ {asof}.xlsx")


def addr_full(r):
    parts = [r[ix["เลขที่"]], r[ix["แขวง/ตำบล/อำเภอ"]], r[ix["จังหวัด"]], r[ix["รหัส ปณ."]]]
    return " ".join(str(p).strip() for p in parts if p is not None and str(p).strip())


def conv(kind, v):
    if v is None:
        return None
    if kind == "int":
        try:
            return int(str(v).strip())
        except (ValueError, TypeError):
            return v
    if kind == "float":
        return float(v)
    if kind == "date":
        return dt.datetime(v.year, v.month, v.day) if hasattr(v, "year") else v
    if kind == "dateBE":
        return dt.datetime(v.year + BE, v.month, v.day) if hasattr(v, "year") else v
    return str(v)


wb = openpyxl.load_workbook(TEMPLATE)
ws = wb["ข้อมูล"]
FIRST = 3
style = {c: dict(number_format=ws[f"{c}{FIRST}"].number_format,
                 font=copy(ws[f"{c}{FIRST}"].font),
                 alignment=copy(ws[f"{c}{FIRST}"].alignment),
                 border=copy(ws[f"{c}{FIRST}"].border),
                 fill=copy(ws[f"{c}{FIRST}"].fill)) for c, *_ in MAP}

if ws.max_row >= FIRST:
    ws.delete_rows(FIRST, ws.max_row - FIRST + 1)

for n, r in enumerate(rows):
    xr = FIRST + n
    for c, src, kind in MAP:
        cell = ws[f"{c}{xr}"]
        if kind == "bahttext":
            cell.value = f"=BAHTTEXT({src}{xr})"
        elif src == "__ADDRESS_FULL__":
            cell.value = addr_full(r)
        else:
            cell.value = conv(kind, r[ix[src]])
        st = style[c]
        cell.number_format = st["number_format"]
        cell.font = copy(st["font"])
        cell.alignment = copy(st["alignment"])
        cell.border = copy(st["border"])
        cell.fill = copy(st["fill"])

if "Sheet1" in wb.sheetnames and wb["Sheet1"].max_row >= 2:
    wb["Sheet1"].delete_rows(2, wb["Sheet1"].max_row - 1)

CHECK = ["สัญญาเลขที่", "ชื่อกรณี", "รอบ", "รหัสลูกค้า",
         "ค้างชำระงวดแรก", "วันครบกำหนดงวดแรกที่ค้าง",
         "ค้างชำระงวดสุดท้าย", "วันครบกำหนดงวดสุดท้ายที่ค้าง",
         "งวดที่ครบ OD 6", "รวมจำนวนงวดที่ค้าง",
         "ค่าเช่าซื้อที่ค้างชำระ", "ค่าเบี้ยปรับชำระล่าช้า",
         "ค่าติดตามทวงถามหนี้", "รวมเป็นเงินทั้งสิ้น"] + [c for c in cols if c.startswith("_")]
ck = wb.create_sheet("ตรวจสอบ")
ck.append(CHECK)
for i in range(1, len(CHECK) + 1):
    ck.cell(1, i).font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    ck.cell(1, i).fill = openpyxl.styles.PatternFill("solid", fgColor="ED7D31")
    ck.column_dimensions[get_column_letter(i)].width = 24
ck.freeze_panes = "B2"
ck.auto_filter.ref = f"A1:{get_column_letter(len(CHECK))}1"
for n, r in enumerate(rows, start=2):
    for i, name in enumerate(CHECK, start=1):
        ck.cell(n, i, r[ix[name]])

# บันทึกพารามิเตอร์ที่ใช้ ไว้ตรวจย้อนหลัง
info = wb.create_sheet("พารามิเตอร์")
for row in [["ดึงข้อมูลเมื่อ", dt.datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["รอบ", label],
            ["รอบที่เอา", "ทุกรอบ" if rnd_param == 0 else f"วันที่ {rnd_param}"],
            ["นับงวดค้างถึงวันที่", str(asof)],
            ["เส้นแบ่งกรณี 1 / 2", str(cut)],
            ["ต้องค้างอย่างน้อย (งวด)", args.min_od],
            ["เอาเฉพาะที่ครบ OD6 เดือนนี้", "ไม่ (เอาทั้งหมด)" if args.all_od6 else "ใช่"],
            ["แหล่งที่ใช้คัด", "CUSTOMER_CARD (ไม่ใช้ COLLECTION_OD)"],
            ["จำนวนสัญญา", len(rows)],
            ["กรณี 1 · OD6 ปกติ", n1],
            ["กรณี 2 · สัญญาหมดอายุ", n2],
            ["ค่าปรับ/ค่าติดตาม", "คิดตาม K2 - Fee Policy"]]:
    info.append(row)
info.column_dimensions["A"].width = 28
info.column_dimensions["B"].width = 24

try:
    wb.save(out)
except PermissionError:
    out = out.with_name(f"{out.stem} ({dt.datetime.now():%H%M}){out.suffix}")
    wb.save(out)
    print(f"  (ไฟล์เดิมเปิดค้างอยู่ จึงเซฟเป็นชื่อใหม่)")
print(f"\nเขียน {out.resolve()}")

if args.csv:
    import csv
    c = out.with_suffix(".csv")
    with open(c, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(cols)
        w.writerows(rows)
    print(f"เขียน {c.resolve()}")

print("• ชีต 'ข้อมูล'      = ตาม Template ทุกคอลัมน์ พร้อมสูตร =BAHTTEXT (เปิดด้วย Excel เท่านั้น)")
print("• ชีต 'ตรวจสอบ'     = คอลัมน์เสริมไว้ตรวจก่อนออกหนังสือ")
print("• ชีต 'พารามิเตอร์' = บันทึกว่าดึงเมื่อไหร่ นับถึงวันไหน เกณฑ์กี่งวด")
print("⚠️ ไฟล์นี้มี PII — ห้าม commit เข้า git")
